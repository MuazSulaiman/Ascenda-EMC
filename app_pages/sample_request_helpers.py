# app_pages/sample_request_helpers.py
"""
Core workflow data-layer for the Sample Request module — mirrors
app_pages/quotation_helpers.py's non-money machinery (guarded state
transitions, revision snapshots, status-event timeline, shared render
helpers) against the sample_requests family of tables. There is
deliberately no money math anywhere in this file: no unit_price,
discount_pct, VAT, or totals — quantity is a plain int and lines carry
only product_id / quantity. delivery_date is a single request-level field
(not per line) alongside request_date.
"""
from urllib.parse import quote_plus

import pandas as pd
import streamlit as st
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError

from db import engine
from db_ops import query_df, query_scalar, exec_sql
from app_pages.change_request_helpers import _norm, _sql_val
from app_pages.notification_helpers import notify_role, notify_users
from ui import compare_row, status_badge, visit_card
from utils import to_local, to_local_str, _local_now


# ─────────────────────────────────────────────────────────────────────────────
# Authorization / duplicate-reference guards
# ─────────────────────────────────────────────────────────────────────────────

def _require_role(actor_uid: int, allowed_roles: set[str]) -> str:
    """Re-query the actor's current role from the DB. Raises ValueError if not allowed."""
    role = query_scalar("SELECT role FROM users WHERE user_id = :uid", {"uid": actor_uid})
    role = _norm(role)
    if role not in allowed_roles:
        raise ValueError(f"Role '{role}' is not permitted to perform this action.")
    return role


def _odoo_reference_exists(ref: str) -> bool:
    ref = _norm(ref)
    if not ref:
        return False
    row = query_scalar(
        "SELECT 1 FROM sample_requests WHERE odoo_reference = :ref", {"ref": ref}
    )
    return row is not None


# ─────────────────────────────────────────────────────────────────────────────
# Atomic sample-request-number assignment
# ─────────────────────────────────────────────────────────────────────────────

def _next_sample_request_number(conn, year: int) -> str:
    result = conn.execute(
        text("""
            INSERT INTO sample_request_number_counters (year, last_seq) VALUES (:yr, 1)
            ON CONFLICT (year) DO UPDATE SET last_seq = sample_request_number_counters.last_seq + 1
            RETURNING last_seq
        """),
        {"yr": year},
    )
    seq = result.scalar_one()
    return f"SDR-{year}-{seq:06d}"


# ─────────────────────────────────────────────────────────────────────────────
# Read helpers
# ─────────────────────────────────────────────────────────────────────────────

def _load_sample_request_header(sample_request_id: int) -> dict | None:
    df = query_df(
        "SELECT * FROM sample_requests WHERE sample_request_id = :id", {"id": sample_request_id}
    )
    return df.iloc[0].to_dict() if not df.empty else None


def _load_sample_request_lines(sample_request_id: int) -> pd.DataFrame:
    return query_df(
        "SELECT * FROM sample_request_lines WHERE sample_request_id = :id ORDER BY line_no",
        {"id": sample_request_id},
    )


def _load_status_events(sample_request_id: int) -> pd.DataFrame:
    return query_df(
        "SELECT * FROM sample_request_status_events WHERE sample_request_id = :id ORDER BY at_utc",
        {"id": sample_request_id},
    )


def _load_revisions(sample_request_id: int) -> pd.DataFrame:
    return query_df(
        "SELECT * FROM sample_request_revisions WHERE sample_request_id = :id ORDER BY revision_no",
        {"id": sample_request_id},
    )


def _load_revision_lines(revision_id: int) -> pd.DataFrame:
    return query_df(
        "SELECT * FROM sample_request_revision_lines WHERE revision_id = :id ORDER BY line_no",
        {"id": revision_id},
    )


# ─────────────────────────────────────────────────────────────────────────────
# Guarded state-transition functions
#
# Idiom matched from app_pages/quotation_helpers.py: a single `engine.begin()`
# block performs one guarded `UPDATE ... WHERE sample_request_id = :sid AND
# status = '<expected>' [...]`. Where the transaction needs to fail loudly on
# a stale/foreign row, we raise ValueError *inside* the `with engine.begin()`
# block (which rolls the whole transaction back) and catch it in an outer
# try/except that converts it to `(False, str(e))`. Manager/coordinator
# functions call `_require_role(...)` as the very first line, before the
# transaction opens, and let a role failure propagate as a raised ValueError
# (defense-in-depth below the page-level role gate — this is not expected to
# be hit through the normal UI, so it is not converted to a soft return).
# ─────────────────────────────────────────────────────────────────────────────

def submit_sample_request(header: dict, lines: list[dict], actor_uid: int) -> tuple[int, str]:
    """
    Insert a brand-new sample request: header (status='IN_REVIEW', version=0),
    lines, revision #1 (+ revision lines), and a SUBMITTED status event, all
    inside one atomic transaction. Pure insert — no status guard needed.
    Returns (sample_request_id, request_number).

    request_date is captured here, server-side, as the org's local "today"
    (never taken from the caller) — the rep does not choose it, matching the
    product decision that this field records when the request was actually
    made, not something editable at submission time.
    """
    request_date = _local_now().date()
    year = request_date.year

    with engine.begin() as conn:
        request_number = _next_sample_request_number(conn, year)

        sample_request_id = conn.execute(
            text("""
                INSERT INTO sample_requests
                    (request_number, customer_id, rep_user_id, request_date, delivery_date,
                     remarks, status, version, submitted_by)
                VALUES
                    (:request_number, :customer_id, :actor_uid, :request_date, :delivery_date,
                     :remarks, 'IN_REVIEW', 0, :actor_uid)
                RETURNING sample_request_id
            """),
            {
                "request_number": request_number,
                "customer_id": header.get("customer_id"),
                "actor_uid": actor_uid,
                "request_date": request_date,
                "delivery_date": header.get("delivery_date"),
                "remarks": _norm(header.get("remarks")) or None,
            },
        ).scalar_one()

        for i, line in enumerate(lines, start=1):
            conn.execute(
                text("""
                    INSERT INTO sample_request_lines
                        (sample_request_id, line_no, product_id, quantity)
                    VALUES (:sid, :line_no, :product_id, :quantity)
                """),
                {
                    "sid": sample_request_id,
                    "line_no": i,
                    "product_id": line["product_id"],
                    "quantity": int(line["quantity"]),
                },
            )

        revision_id = conn.execute(
            text("""
                INSERT INTO sample_request_revisions
                    (sample_request_id, revision_no, created_by, customer_id, request_date,
                     delivery_date, remarks)
                VALUES
                    (:sid, 1, :actor_uid, :customer_id, :request_date, :delivery_date, :remarks)
                RETURNING revision_id
            """),
            {
                "sid": sample_request_id,
                "actor_uid": actor_uid,
                "customer_id": header.get("customer_id"),
                "request_date": request_date,
                "delivery_date": header.get("delivery_date"),
                "remarks": _norm(header.get("remarks")) or None,
            },
        ).scalar_one()

        for i, line in enumerate(lines, start=1):
            item_row = conn.execute(
                text("SELECT article_number, description FROM items WHERE product_id = :pid"),
                {"pid": line["product_id"]},
            ).mappings().one_or_none()
            conn.execute(
                text("""
                    INSERT INTO sample_request_revision_lines
                        (revision_id, line_no, product_id, article_number_snapshot, description_snapshot,
                         quantity)
                    VALUES
                        (:rid, :line_no, :product_id, :article_number, :description,
                         :quantity)
                """),
                {
                    "rid": revision_id,
                    "line_no": i,
                    "product_id": line["product_id"],
                    "article_number": item_row["article_number"] if item_row else None,
                    "description": item_row["description"] if item_row else None,
                    "quantity": int(line["quantity"]),
                },
            )

        conn.execute(
            text("""
                INSERT INTO sample_request_status_events
                    (sample_request_id, event_type, actor_user_id, from_status, to_status, revision_id)
                VALUES (:sid, 'SUBMITTED', :actor_uid, NULL, 'IN_REVIEW', :rid)
            """),
            {"sid": sample_request_id, "actor_uid": actor_uid, "rid": revision_id},
        )

        notify_role(
            conn, ["sales manager", "admin"], exclude_user_id=actor_uid,
            category="sample_request", event_type="SUBMITTED",
            title=f"New sample request {request_number} awaiting review",
            link_page="Review Sample Requests", link_params={"sample_request_id": sample_request_id},
            actor_user_id=actor_uid,
        )

    return int(sample_request_id), request_number


def resubmit_sample_request(
    sample_request_id: int, header: dict, lines: list[dict], actor_uid: int, expected_version: int
) -> tuple[bool, str | None]:
    """
    Rep resubmits an EDIT_REQUESTED sample request. Single guarded UPDATE
    combines the version bump, status guard, and rep-ownership guard in one
    WHERE clause — the only mechanism that decides whether this call wins the
    race. Everything (line replace, header update, new revision snapshot,
    event) happens in the same transaction; any failure rolls all of it back.
    request_date is never touched here — it's set once at submit_sample_request
    time and stays fixed for the life of the request, including every resubmit.
    """
    try:
        with engine.begin() as conn:
            result = conn.execute(
                text("""
                    UPDATE sample_requests
                    SET version = version + 1, status = 'IN_REVIEW'
                    WHERE sample_request_id = :sid AND status = 'EDIT_REQUESTED'
                      AND version = :expected_version AND rep_user_id = :actor_uid
                    RETURNING version, request_number, request_date
                """),
                {"sid": sample_request_id, "expected_version": expected_version, "actor_uid": actor_uid},
            )
            row = result.fetchone()
            if row is None:
                raise ValueError("This sample request was already changed elsewhere. Refresh and try again.")
            new_version = row[0]
            request_number = row[1]
            request_date = row[2]  # immutable — set once at submission, never editable on resubmit

            conn.execute(
                text("DELETE FROM sample_request_lines WHERE sample_request_id = :sid"),
                {"sid": sample_request_id},
            )

            for i, line in enumerate(lines, start=1):
                conn.execute(
                    text("""
                        INSERT INTO sample_request_lines
                            (sample_request_id, line_no, product_id, quantity)
                        VALUES (:sid, :line_no, :product_id, :quantity)
                    """),
                    {
                        "sid": sample_request_id,
                        "line_no": i,
                        "product_id": line["product_id"],
                        "quantity": int(line["quantity"]),
                    },
                )

            conn.execute(
                text("""
                    UPDATE sample_requests
                    SET customer_id = :customer_id, delivery_date = :delivery_date, remarks = :remarks
                    WHERE sample_request_id = :sid
                """),
                {
                    "sid": sample_request_id,
                    "customer_id": header.get("customer_id"),
                    "delivery_date": header.get("delivery_date"),
                    "remarks": _norm(header.get("remarks")) or None,
                },
            )

            revision_no = new_version + 1
            revision_id = conn.execute(
                text("""
                    INSERT INTO sample_request_revisions
                        (sample_request_id, revision_no, created_by, customer_id, request_date,
                         delivery_date, remarks)
                    VALUES
                        (:sid, :revision_no, :actor_uid, :customer_id, :request_date,
                         :delivery_date, :remarks)
                    RETURNING revision_id
                """),
                {
                    "sid": sample_request_id,
                    "revision_no": revision_no,
                    "actor_uid": actor_uid,
                    "customer_id": header.get("customer_id"),
                    "request_date": request_date,
                    "delivery_date": header.get("delivery_date"),
                    "remarks": _norm(header.get("remarks")) or None,
                },
            ).scalar_one()

            for i, line in enumerate(lines, start=1):
                item_row = conn.execute(
                    text("SELECT article_number, description FROM items WHERE product_id = :pid"),
                    {"pid": line["product_id"]},
                ).mappings().one_or_none()
                conn.execute(
                    text("""
                        INSERT INTO sample_request_revision_lines
                            (revision_id, line_no, product_id, article_number_snapshot, description_snapshot,
                             quantity)
                        VALUES
                            (:rid, :line_no, :product_id, :article_number, :description,
                             :quantity)
                    """),
                    {
                        "rid": revision_id,
                        "line_no": i,
                        "product_id": line["product_id"],
                        "article_number": item_row["article_number"] if item_row else None,
                        "description": item_row["description"] if item_row else None,
                        "quantity": int(line["quantity"]),
                    },
                )

            conn.execute(
                text("""
                    INSERT INTO sample_request_status_events
                        (sample_request_id, event_type, actor_user_id, from_status, to_status, revision_id)
                    VALUES (:sid, 'RESUBMITTED', :actor_uid, 'EDIT_REQUESTED', 'IN_REVIEW', :rid)
                """),
                {"sid": sample_request_id, "actor_uid": actor_uid, "rid": revision_id},
            )

            notify_role(
                conn, ["sales manager", "admin"], exclude_user_id=actor_uid,
                category="sample_request", event_type="RESUBMITTED",
                title=f"Sample request {request_number} resubmitted for review",
                link_page="Review Sample Requests", link_params={"sample_request_id": sample_request_id},
                actor_user_id=actor_uid,
            )
        return True, None
    except Exception as e:
        return False, str(e)


def manager_approve(sample_request_id: int, manager_uid: int) -> tuple[bool, str | None]:
    role = _require_role(manager_uid, {"sales manager", "admin"})
    self_review_clause = "" if role == "admin" else "AND rep_user_id != :actor_uid"
    try:
        with engine.begin() as conn:
            result = conn.execute(
                text(f"""
                    UPDATE sample_requests
                    SET status = 'APPROVED', manager_user_id = :actor_uid, manager_decided_at = NOW()
                    WHERE sample_request_id = :sid AND status = 'IN_REVIEW' {self_review_clause}
                    RETURNING sample_request_id, rep_user_id, request_number
                """),
                {"sid": sample_request_id, "actor_uid": manager_uid},
            )
            row = result.fetchone()
            if row is None:
                raise ValueError(
                    "This sample request was already resolved elsewhere, or you cannot review your "
                    "own submission. Refresh and try again."
                )
            rep_user_id, request_number = row[1], row[2]
            conn.execute(
                text("""
                    INSERT INTO sample_request_status_events
                        (sample_request_id, event_type, actor_user_id, from_status, to_status, revision_id)
                    VALUES (:sid, 'APPROVED', :actor_uid, 'IN_REVIEW', 'APPROVED', NULL)
                """),
                {"sid": sample_request_id, "actor_uid": manager_uid},
            )

            notify_users(
                conn, [rep_user_id], category="sample_request", event_type="APPROVED",
                title=f"Your sample request {request_number} was approved",
                link_page="Sample Requests", link_params={"sample_request_id": sample_request_id},
                actor_user_id=manager_uid,
            )
            notify_role(
                conn, ["sales coordinator", "admin"], category="sample_request", event_type="APPROVED",
                title=f"Sample request {request_number} approved, ready for handoff",
                link_page="Sample Request Handoff (Odoo)", link_params={"sample_request_id": sample_request_id},
                actor_user_id=manager_uid,
            )
        return True, None
    except Exception as e:
        return False, str(e)


def manager_reject(sample_request_id: int, manager_uid: int, reason: str) -> tuple[bool, str | None]:
    role = _require_role(manager_uid, {"sales manager", "admin"})
    if not (reason or "").strip():
        return False, "A rejection reason is required."
    self_review_clause = "" if role == "admin" else "AND rep_user_id != :actor_uid"
    try:
        with engine.begin() as conn:
            result = conn.execute(
                text(f"""
                    UPDATE sample_requests
                    SET status = 'REJECTED', manager_user_id = :actor_uid, manager_decided_at = NOW(),
                        manager_comment = :comment
                    WHERE sample_request_id = :sid AND status = 'IN_REVIEW' {self_review_clause}
                    RETURNING sample_request_id, rep_user_id, request_number
                """),
                {"sid": sample_request_id, "actor_uid": manager_uid, "comment": reason.strip()},
            )
            row = result.fetchone()
            if row is None:
                raise ValueError(
                    "This sample request was already resolved elsewhere, or you cannot review your "
                    "own submission. Refresh and try again."
                )
            rep_user_id, request_number = row[1], row[2]
            conn.execute(
                text("""
                    INSERT INTO sample_request_status_events
                        (sample_request_id, event_type, actor_user_id, from_status, to_status, comment, revision_id)
                    VALUES (:sid, 'REJECTED', :actor_uid, 'IN_REVIEW', 'REJECTED', :comment, NULL)
                """),
                {"sid": sample_request_id, "actor_uid": manager_uid, "comment": reason.strip()},
            )

            notify_users(
                conn, [rep_user_id], category="sample_request", event_type="REJECTED",
                title=f"Your sample request {request_number} was rejected",
                link_page="Sample Requests", link_params={"sample_request_id": sample_request_id},
                actor_user_id=manager_uid,
            )
        return True, None
    except Exception as e:
        return False, str(e)


def manager_request_edit(sample_request_id: int, manager_uid: int, comment: str) -> tuple[bool, str | None]:
    role = _require_role(manager_uid, {"sales manager", "admin"})
    if not (comment or "").strip():
        return False, "A comment explaining the requested edit is required."
    self_review_clause = "" if role == "admin" else "AND rep_user_id != :actor_uid"
    try:
        with engine.begin() as conn:
            result = conn.execute(
                text(f"""
                    UPDATE sample_requests
                    SET status = 'EDIT_REQUESTED', manager_user_id = :actor_uid, manager_decided_at = NOW(),
                        manager_comment = :comment
                    WHERE sample_request_id = :sid AND status = 'IN_REVIEW' {self_review_clause}
                    RETURNING sample_request_id, rep_user_id, request_number
                """),
                {"sid": sample_request_id, "actor_uid": manager_uid, "comment": comment.strip()},
            )
            row = result.fetchone()
            if row is None:
                raise ValueError(
                    "This sample request was already resolved elsewhere, or you cannot review your "
                    "own submission. Refresh and try again."
                )
            rep_user_id, request_number = row[1], row[2]
            conn.execute(
                text("""
                    INSERT INTO sample_request_status_events
                        (sample_request_id, event_type, actor_user_id, from_status, to_status, comment, revision_id)
                    VALUES (:sid, 'EDIT_REQUESTED', :actor_uid, 'IN_REVIEW', 'EDIT_REQUESTED', :comment, NULL)
                """),
                {"sid": sample_request_id, "actor_uid": manager_uid, "comment": comment.strip()},
            )

            notify_users(
                conn, [rep_user_id], category="sample_request", event_type="EDIT_REQUESTED",
                title=f"Edit requested on your sample request {request_number}",
                link_page="Sample Requests", link_params={"sample_request_id": sample_request_id},
                actor_user_id=manager_uid,
            )
        return True, None
    except Exception as e:
        return False, str(e)


def manager_return_for_revision(sample_request_id: int, manager_uid: int, reason: str) -> tuple[bool, str | None]:
    role = _require_role(manager_uid, {"sales manager", "admin"})
    if not (reason or "").strip():
        return False, "A reason for returning this sample request is required."
    self_review_clause = "" if role == "admin" else "AND rep_user_id != :actor_uid"
    try:
        with engine.begin() as conn:
            result = conn.execute(
                text(f"""
                    UPDATE sample_requests
                    SET status = 'EDIT_REQUESTED', manager_user_id = :actor_uid, manager_decided_at = NOW(),
                        manager_comment = :comment
                    WHERE sample_request_id = :sid AND status = 'APPROVED' {self_review_clause}
                    RETURNING sample_request_id, rep_user_id, request_number
                """),
                {"sid": sample_request_id, "actor_uid": manager_uid, "comment": reason.strip()},
            )
            row = result.fetchone()
            if row is None:
                raise ValueError(
                    "This sample request was already resolved elsewhere, or you cannot review your "
                    "own submission. Refresh and try again."
                )
            rep_user_id, request_number = row[1], row[2]
            conn.execute(
                text("""
                    INSERT INTO sample_request_status_events
                        (sample_request_id, event_type, actor_user_id, from_status, to_status, comment, revision_id)
                    VALUES (:sid, 'RETURNED_FOR_REVISION', :actor_uid, 'APPROVED', 'EDIT_REQUESTED', :comment, NULL)
                """),
                {"sid": sample_request_id, "actor_uid": manager_uid, "comment": reason.strip()},
            )

            notify_users(
                conn, [rep_user_id], category="sample_request", event_type="RETURNED_FOR_REVISION",
                title=f"Your sample request {request_number} was returned for revision",
                link_page="Sample Requests", link_params={"sample_request_id": sample_request_id},
                actor_user_id=manager_uid,
            )
        return True, None
    except Exception as e:
        return False, str(e)


def withdraw_sample_request(sample_request_id: int, rep_uid: int, reason: str) -> tuple[bool, str | None]:
    """
    Owning rep withdraws from IN_REVIEW or EDIT_REQUESTED. The guard allows two
    possible prior statuses, so the accurate `from_status` for the audit event
    is captured with a `FOR UPDATE` CTE folded into the *same* guarded UPDATE
    statement (Postgres evaluates the CTE against the pre-UPDATE snapshot) —
    this is still exactly one guarded UPDATE, not a separate SELECT-then-UPDATE:
    if the WHERE guard doesn't match, zero rows come back and nothing else in
    the transaction runs. No notification (matches quotations' withdraw_quotation).
    """
    if not (reason or "").strip():
        return False, "A withdrawal reason is required."
    try:
        with engine.begin() as conn:
            result = conn.execute(
                text("""
                    WITH prior AS (
                        SELECT status FROM sample_requests WHERE sample_request_id = :sid FOR UPDATE
                    )
                    UPDATE sample_requests
                    SET status = 'WITHDRAWN', withdrawn_at = NOW(), withdrawn_by = :actor_uid,
                        withdrawal_reason = :reason
                    FROM prior
                    WHERE sample_requests.sample_request_id = :sid
                      AND prior.status IN ('IN_REVIEW', 'EDIT_REQUESTED')
                      AND sample_requests.rep_user_id = :actor_uid
                    RETURNING sample_requests.sample_request_id, prior.status AS prior_status
                """),
                {"sid": sample_request_id, "actor_uid": rep_uid, "reason": reason.strip()},
            )
            row = result.fetchone()
            if row is None:
                raise ValueError(
                    "This sample request could not be withdrawn — it may have already been resolved, "
                    "or it does not belong to you. Refresh and try again."
                )
            prior_status = row[1]

            conn.execute(
                text("""
                    INSERT INTO sample_request_status_events
                        (sample_request_id, event_type, actor_user_id, from_status, to_status, comment, revision_id)
                    VALUES (:sid, 'WITHDRAWN', :actor_uid, :from_status, 'WITHDRAWN', :reason, NULL)
                """),
                {"sid": sample_request_id, "actor_uid": rep_uid, "from_status": prior_status, "reason": reason.strip()},
            )
        return True, None
    except Exception as e:
        return False, str(e)


def coordinator_mark_done(
    sample_request_id: int, coordinator_uid: int, odoo_reference: str | None, note: str | None
) -> tuple[bool, str | None]:
    """
    Coordinator marks an APPROVED sample request DONE. `_odoo_reference_exists()`
    is the primary defense (caller/UI should already have checked it and
    disabled the button); the DB-level guarded UPDATE plus the unique partial
    index `idx_sr_odoo_ref_unique` is the backstop for the true concurrent-race
    case, caught here as `IntegrityError` and converted to a soft `(False, ...)`.

    Notification recipients mirror quotation_helpers.coordinator_mark_done
    exactly: two separate notify_users calls (rep, then manager) rather than
    one combined call, each with its own link_page (rep -> "Sample Requests",
    manager -> "Review Sample Requests"). manager_user_id is not None-guarded
    because it structurally cannot be NULL here — this guarded UPDATE only
    ever matches a row whose status is 'APPROVED', and the only path to
    'APPROVED' (manager_approve) always sets manager_user_id in the same
    transaction — exactly the same invariant quotations relies on.
    """
    _require_role(coordinator_uid, {"sales coordinator", "admin"})
    ref = _norm(odoo_reference) or None
    note_val = _norm(note) or None

    if ref and _odoo_reference_exists(ref):
        return False, f"Odoo reference '{ref}' is already used by another sample request."

    try:
        with engine.begin() as conn:
            result = conn.execute(
                text("""
                    UPDATE sample_requests
                    SET status = 'DONE', coordinator_user_id = :actor_uid, coordinator_done_at = NOW(),
                        coordinator_note = :note, odoo_reference = :ref
                    WHERE sample_request_id = :sid AND status = 'APPROVED'
                    RETURNING sample_request_id, rep_user_id, manager_user_id, request_number
                """),
                {"sid": sample_request_id, "actor_uid": coordinator_uid, "note": note_val, "ref": ref},
            )
            row = result.fetchone()
            if row is None:
                raise ValueError("This sample request was already resolved elsewhere. Refresh and try again.")
            rep_user_id, manager_user_id, request_number = row[1], row[2], row[3]
            conn.execute(
                text("""
                    INSERT INTO sample_request_status_events
                        (sample_request_id, event_type, actor_user_id, from_status, to_status, comment, revision_id)
                    VALUES (:sid, 'MARKED_DONE', :actor_uid, 'APPROVED', 'DONE', :note, NULL)
                """),
                {"sid": sample_request_id, "actor_uid": coordinator_uid, "note": note_val},
            )

            notify_users(
                conn, [rep_user_id], category="sample_request", event_type="MARKED_DONE",
                title=f"Sample request {request_number} marked Done",
                link_page="Sample Requests", link_params={"sample_request_id": sample_request_id},
                actor_user_id=coordinator_uid,
            )
            notify_users(
                conn, [manager_user_id], category="sample_request", event_type="MARKED_DONE",
                title=f"Sample request {request_number} marked Done",
                link_page="Review Sample Requests", link_params={"sample_request_id": sample_request_id},
                actor_user_id=coordinator_uid,
            )
        return True, None
    except IntegrityError:
        return False, "That Odoo reference was just used by another sample request."
    except Exception as e:
        return False, str(e)


# ─────────────────────────────────────────────────────────────────────────────
# Shared rendering helpers
# ─────────────────────────────────────────────────────────────────────────────

def render_search_date_filter(ns: str, *, search_placeholder: str = "Search…", show_date: bool = True):
    """
    Shared search + optional date-range controls, matching the pattern already
    used in quotation_helpers.render_search_date_filter.
    Returns (search_q, date_from, date_to) for use with filter_sample_requests_df.
    """
    search_q = st.text_input(
        "", placeholder=search_placeholder, key=f"{ns}_search", label_visibility="collapsed"
    )
    date_from = date_to = None
    if show_date:
        c1, c2 = st.columns(2)
        with c1:
            date_from = st.date_input("From", value=None, key=f"{ns}_date_from")
        with c2:
            date_to = st.date_input("To", value=None, key=f"{ns}_date_to")
    return search_q, date_from, date_to


def filter_sample_requests_df(
    df: pd.DataFrame,
    *,
    search_q: str = "",
    date_from=None,
    date_to=None,
    date_col: str = "submitted_at",
    text_cols: tuple = ("request_number", "account_name"),
) -> pd.DataFrame:
    """Client-side filter of an already-loaded sample-requests DataFrame — no new queries."""
    if df.empty:
        return df
    out = df

    if (date_from or date_to) and date_col in out.columns:
        flt_date = pd.to_datetime(out[date_col], errors="coerce").dt.date
        if date_from:
            out = out[flt_date >= date_from]
            flt_date = flt_date[out.index]
        if date_to:
            out = out[flt_date <= date_to]

    q = (search_q or "").strip().lower()
    if q:
        mask = pd.Series(False, index=out.index)
        for col in text_cols:
            if col in out.columns:
                mask = mask | out[col].astype(str).str.lower().str.contains(q, na=False, regex=False)
        out = out[mask]

    return out


def sample_request_detail_href(page_name: str, sample_request_id: int) -> str:
    """URL to a single sample request's routed detail view (?page=...&sample_request_id=...&_sid=...)."""
    href = f"?page={quote_plus(page_name)}&sample_request_id={int(sample_request_id)}"
    nav_sid = st.session_state.get("_stored_sid", "")
    if nav_sid:
        href += f"&_sid={nav_sid}"
    return href


def render_sample_request_list(
    ns: str,
    df: pd.DataFrame,
    *,
    page_name: str,
    status_labels: dict,
    status_order: list,
    search_text_cols: tuple = ("request_number", "account_name"),
    date_col: str = "submitted_at",
    search_placeholder: str = "Search…",
    page_size: int = 10,
) -> None:
    """
    Shared paginated list view for a sample-requests queue: search + date
    filter, counted status tabs, and card rows (reusing ui.visit_card, which
    is fully generic) linking to a routed detail view via
    sample_request_detail_href. Mirrors quotation_helpers.render_quotation_list.
    """
    if df.empty:
        st.info("No sample requests to show.")
        return

    search_q, date_from, date_to = render_search_date_filter(ns, search_placeholder=search_placeholder)
    filtered = filter_sample_requests_df(
        df, search_q=search_q, date_from=date_from, date_to=date_to,
        date_col=date_col, text_cols=search_text_cols,
    )

    tab_labels = [f"All  {len(filtered)}"]
    for status in status_order:
        cnt = int((filtered["status"] == status).sum()) if "status" in filtered.columns else 0
        tab_labels.append(f"{status_labels.get(status, status)}  {cnt}")

    chosen_label = st.radio(
        "Status", tab_labels, key=f"{ns}_status_tab", horizontal=True, label_visibility="collapsed",
    )
    chosen = chosen_label.rsplit("  ", 1)[0].strip()

    if chosen == "All":
        visible = filtered
    else:
        inv_labels = {v: k for k, v in status_labels.items()}
        status_key = inv_labels.get(chosen, chosen)
        visible = filtered[filtered["status"] == status_key]

    if visible.empty:
        st.caption("No results — try a different search, date range, or status.")
        return

    total = len(visible)
    filter_key = (search_q, date_from, date_to, chosen_label)
    page_state_key = f"{ns}_page"
    filter_state_key = f"{ns}_filter_key"
    if st.session_state.get(filter_state_key) != filter_key:
        st.session_state[filter_state_key] = filter_key
        st.session_state[page_state_key] = 0

    current_page = st.session_state.get(page_state_key, 0)
    total_pages = max(1, (total + page_size - 1) // page_size)
    current_page = min(current_page, total_pages - 1)

    start_idx = current_page * page_size
    end_idx = min(start_idx + page_size, total)
    page_df = visible.iloc[start_idx:end_idx]

    st.caption(f"Showing {start_idx + 1}–{end_idx} of {total:,} sample request{'s' if total != 1 else ''}")

    cards_html = ""
    for _, row in page_df.iterrows():
        sid = int(row["sample_request_id"])
        status_val = _norm(row.get("status"))
        status_label = status_labels.get(status_val, status_val)
        cards_html += visit_card(
            _norm(row.get("rep_name")),
            to_local(row.get(date_col)),
            _norm(row.get("account_name")) or "—",
            subtitle=_norm(row.get("request_number")),
            status=status_label,
            status_variant=_status_badge_variant(status_val),
            href=sample_request_detail_href(page_name, sid),
        )
    st.markdown(cards_html, unsafe_allow_html=True)

    if total_pages > 1:
        pcol1, pcol2, pcol3 = st.columns([1, 2, 1])
        with pcol1:
            if st.button("← Prev", key=f"{ns}_prev_btn", disabled=current_page <= 0, use_container_width=True):
                st.session_state[page_state_key] = current_page - 1
                st.rerun()
        with pcol2:
            st.markdown(
                f'<p style="text-align:center;color:var(--color-text-subtle);font-size:0.85rem;margin-top:0.4rem;">'
                f'Page {current_page + 1} of {total_pages}</p>',
                unsafe_allow_html=True,
            )
        with pcol3:
            if st.button("Next →", key=f"{ns}_next_btn", disabled=current_page >= total_pages - 1, use_container_width=True):
                st.session_state[page_state_key] = current_page + 1
                st.rerun()


_REVISION_HEADER_FIELDS = [
    ("request_date", "Request Date"),
    ("delivery_date", "Delivery Date"),
    ("remarks", "Remarks"),
]


def render_revision_diff(revision_a: dict, revision_b: dict) -> None:
    """
    Header-field-by-field and line-by-line comparison of two
    sample_request_revisions rows, using the same visual idiom as
    ui.py::compare_row (reused, not reimplemented). No totals row.

    Each dict is a revision row (e.g. from `_load_revisions(...).iloc[i].to_dict()`)
    with an added "lines" key: a list of dicts from
    `_load_revision_lines(revision_id).to_dict("records")`.
    """
    label_a = f"Revision #{revision_a.get('revision_no', '?')}"
    label_b = f"Revision #{revision_b.get('revision_no', '?')}"

    header_rows = []
    for col, label in _REVISION_HEADER_FIELDS:
        old_val = _norm(revision_a.get(col))
        new_val = _norm(revision_b.get(col))
        header_rows.append(compare_row(label, old_val, new_val, changed=(old_val != new_val)))

    lines_a = {ln["line_no"]: ln for ln in (revision_a.get("lines") or [])}
    lines_b = {ln["line_no"]: ln for ln in (revision_b.get("lines") or [])}
    line_rows = []
    for line_no in sorted(set(lines_a) | set(lines_b)):
        la = lines_a.get(line_no)
        lb = lines_b.get(line_no)

        def _line_desc(ln):
            if not ln:
                return "(none)"
            art = _norm(ln.get("article_number_snapshot")) or _norm(ln.get("product_id"))
            return f"{art} — qty {ln.get('quantity')}"

        old_desc = _line_desc(la)
        new_desc = _line_desc(lb)
        line_rows.append(compare_row(f"Line {line_no}", old_desc, new_desc, changed=(old_desc != new_desc)))

    rows_html = "".join(header_rows) + "".join(line_rows)
    st.markdown(
        f"""
        <table style="width:100%;border-collapse:collapse;border:1px solid var(--color-border);
                      border-radius:10px;overflow:hidden;font-size:0.875rem;">
          <thead>
            <tr style="background:var(--color-surface-2);">
              <th style="padding:10px 12px;text-align:left;font-weight:600;color:var(--color-text-muted);
                         border-bottom:1px solid var(--color-border);width:30%;">Field</th>
              <th style="padding:10px 12px;text-align:left;font-weight:600;color:var(--color-text-muted);
                         border-bottom:1px solid var(--color-border);">{label_a}</th>
              <th style="padding:10px 12px;text-align:left;font-weight:600;color:var(--color-text-muted);
                         border-bottom:1px solid var(--color-border);">{label_b}</th>
            </tr>
          </thead>
          <tbody>{rows_html}</tbody>
        </table>
        """,
        unsafe_allow_html=True,
    )


def render_sample_request_detail(sample_request_id: int) -> None:
    """
    Shared read-only sample-request view (header, lines, status timeline, and
    a latest-two-revisions diff when applicable) used identically by every
    sample-request-facing page, so the detail view isn't copy-pasted three
    times. Mirrors quotation_helpers.render_quotation_detail's structure and
    layout, dropping every money/totals section (there is none here).
    """
    header = _load_sample_request_header(sample_request_id)
    if not header:
        st.warning("Sample request not found.")
        return

    lines_df = query_df(
        """
        SELECT srl.line_no, srl.product_id, i.article_number, i.description, i.unit_of_measurement,
               srl.quantity
        FROM sample_request_lines srl
        JOIN items i ON i.product_id = srl.product_id
        WHERE srl.sample_request_id = :id
        ORDER BY srl.line_no
        """,
        {"id": sample_request_id},
    )

    status_val = _norm(header.get("status"))
    st.markdown(
        f"#### {_norm(header.get('request_number'))} &nbsp; "
        + status_badge(status_val, _status_badge_variant(status_val)),
        unsafe_allow_html=True,
    )

    cust_name = query_scalar(
        "SELECT account_name FROM customers WHERE customer_id = :cid", {"cid": header.get("customer_id")}
    )
    rep_name = query_scalar(
        "SELECT name FROM users WHERE user_id = :uid", {"uid": header.get("rep_user_id")}
    )

    col1, col2 = st.columns(2)
    with col1:
        st.write(f"**Customer:** {_norm(cust_name) or '—'}")
        st.write(f"**Sales Rep:** {_norm(rep_name) or '—'}")
    with col2:
        st.write(f"**Request Date:** {header.get('request_date') or '—'}")
        st.write(f"**Delivery Date:** {header.get('delivery_date') or '—'}")
        st.write(f"**Remarks:** {_norm(header.get('remarks')) or '—'}")

    st.markdown("##### Line Items")
    if lines_df.empty:
        st.info("No line items.")
    else:
        display_df = lines_df.copy()
        display_df["unit_of_measurement"] = display_df["unit_of_measurement"].fillna("")
        st.dataframe(
            display_df[
                ["line_no", "article_number", "description", "unit_of_measurement", "quantity"]
            ],
            use_container_width=True,
            hide_index=True,
            column_config={"unit_of_measurement": "Unit"},
        )

    if status_val in ("APPROVED", "DONE"):
        odoo_ref = _norm(header.get("odoo_reference"))
        coord_note = _norm(header.get("coordinator_note"))
        st.markdown("##### Handoff to Odoo")
        if odoo_ref or coord_note:
            if odoo_ref:
                st.write(f"**Odoo Reference:** {odoo_ref}")
            if coord_note:
                st.write(f"**Coordinator Note:** {coord_note}")
        else:
            st.caption("No Odoo reference recorded yet.")

    revisions_df = _load_revisions(sample_request_id)
    if len(revisions_df) >= 2:
        st.markdown("##### Revision Comparison (latest two)")
        rev_a = revisions_df.iloc[-2].to_dict()
        rev_b = revisions_df.iloc[-1].to_dict()
        rev_a["lines"] = _load_revision_lines(int(rev_a["revision_id"])).to_dict("records")
        rev_b["lines"] = _load_revision_lines(int(rev_b["revision_id"])).to_dict("records")
        render_revision_diff(rev_a, rev_b)

    st.markdown("##### Status Timeline")
    events_df = _load_status_events(sample_request_id)
    if events_df.empty:
        st.caption("No status events recorded.")
        return

    name_map: dict = {}
    actor_ids = [int(x) for x in events_df["actor_user_id"].dropna().unique().tolist()]
    if actor_ids:
        names_df = query_df("SELECT user_id, name FROM users WHERE user_id = ANY(:ids)", {"ids": actor_ids})
        if not names_df.empty:
            name_map = dict(zip(names_df["user_id"], names_df["name"]))

    for _, ev in events_df.iterrows():
        actor_name = name_map.get(ev.get("actor_user_id"), "—")
        at_str = to_local_str(ev.get("at_utc"))
        comment = _norm(ev.get("comment"))
        from_s = _norm(ev.get("from_status")) or "—"
        to_s = _norm(ev.get("to_status"))
        event_type = _norm(ev.get("event_type"))
        badge_html = status_badge(f"{event_type}: {from_s} → {to_s}", _status_badge_variant(to_s))

        st.markdown(
            f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:4px;">'
            f'{badge_html}'
            f'<span style="font-size:0.8rem;color:var(--color-text-subtle);">{_norm(actor_name)} · {at_str}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )
        if comment:
            st.caption(f'"{comment}"')


def _status_badge_variant(status: str) -> str:
    return {
        "IN_REVIEW": "info",
        "EDIT_REQUESTED": "warning",
        "REJECTED": "danger",
        "APPROVED": "success",
        "DONE": "success",
        "WITHDRAWN": "neutral",
    }.get(status, "neutral")


# ─────────────────────────────────────────────────────────────────────────────
# Warehouse pick-sheet export
# ─────────────────────────────────────────────────────────────────────────────

_PICK_SHEET_COLUMNS = 8  # SI.NO .. Remarks


def _pick_sheet_fmt_date(val) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)) or pd.isna(val):
        return "—"
    if hasattr(val, "strftime"):
        return val.strftime("%d %b %Y")
    return str(val)


def generate_sample_pick_sheet(sample_request_id: int) -> bytes:
    """
    Build a warehouse-facing pick sheet as a styled .xlsx workbook for an
    APPROVED or DONE sample request. There is no quotation-PDF equivalent
    to mirror here (quotations render a PDF via PyMuPDF) — this is a fresh
    workbook via openpyxl. One row per physical unit: a line's quantity is
    exploded into `quantity` separate SI.NO rows (Unit column reading
    "1 of N".."N of N"), leaving Serial/Batch Number, Warehouse Name,
    Warehouse Location, and Remarks blank for the warehouse team to
    hand-fill (those four columns are shaded to flag them as hand-fill
    fields).
    """
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header = _load_sample_request_header(sample_request_id)
    if not header:
        raise ValueError(f"Sample request {sample_request_id} not found.")
    status = _norm(header.get("status"))
    if status not in ("APPROVED", "DONE"):
        raise ValueError(
            f"Pick sheet can only be generated for APPROVED or DONE sample requests "
            f"(current status: {status})."
        )

    lines_df = _load_sample_request_lines(sample_request_id)

    account_name = query_scalar(
        "SELECT account_name FROM customers WHERE customer_id = :cid", {"cid": header.get("customer_id")}
    )

    item_lookup: dict = {}
    product_ids = lines_df["product_id"].dropna().unique().tolist() if not lines_df.empty else []
    if product_ids:
        items_df = query_df(
            "SELECT product_id, article_number, description FROM items WHERE product_id = ANY(:pids)",
            {"pids": product_ids},
        )
        item_lookup = {
            row["product_id"]: (row["article_number"], row["description"])
            for _, row in items_df.iterrows()
        }

    # ── styles ──────────────────────────────────────────────────────────────
    NAVY = "1F3864"
    STEEL = "2E5395"
    BAND = "F2F5FA"
    HANDFILL = "FFF6D9"
    THIN = Side(style="thin", color="B7C3D6")
    BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    title_font = Font(bold=True, size=15, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor=NAVY)
    label_font = Font(bold=True, size=10, color="1F3864")
    value_font = Font(size=10, color="000000")
    value_font_strong = Font(bold=True, size=11, color="000000")
    info_fill = PatternFill("solid", fgColor=BAND)
    col_header_font = Font(bold=True, size=10.5, color="FFFFFF")
    col_header_fill = PatternFill("solid", fgColor=STEEL)
    handfill_fill = PatternFill("solid", fgColor=HANDFILL)
    band_fill = PatternFill("solid", fgColor=BAND)
    footer_font = Font(italic=True, size=8.5, color="6B7280")

    wb = Workbook()
    ws = wb.active
    ws.title = "Warehouse Pick Sheet"

    for i, width in enumerate([9, 14, 42, 12, 20, 18, 22, 28], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── title band ──────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_PICK_SHEET_COLUMNS)
    title_cell = ws.cell(row=1, column=1, value="WAREHOUSE PICK SHEET")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    for col in range(1, _PICK_SHEET_COLUMNS + 1):
        ws.cell(row=1, column=col).fill = title_fill

    ws.row_dimensions[2].height = 6  # thin spacer

    # ── info block: two label/value pairs per row ──────────────────────────
    def _info_row(row: int, left_label: str, left_val: str, right_label: str, right_val: str, strong=False):
        ws.cell(row=row, column=1, value=left_label).font = label_font
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        c = ws.cell(row=row, column=2, value=left_val)
        c.font = value_font_strong if strong else value_font

        ws.cell(row=row, column=4, value=right_label).font = label_font
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=_PICK_SHEET_COLUMNS)
        c2 = ws.cell(row=row, column=5, value=right_val)
        c2.font = value_font_strong if strong else value_font

        for col in range(1, _PICK_SHEET_COLUMNS + 1):
            ws.cell(row=row, column=col).fill = info_fill

    _info_row(
        3,
        "Request No:", _norm(header.get("request_number")) or "—",
        "Date:", _pick_sheet_fmt_date(header.get("request_date")),
        strong=True,
    )
    _info_row(
        4,
        "Delivery Date:", _pick_sheet_fmt_date(header.get("delivery_date")),
        "Customer:", _norm(account_name) or "—",
        strong=True,
    )
    ws.row_dimensions[5].height = 8  # thin spacer

    # ── table header ─────────────────────────────────────────────────────────
    header_row = 6
    col_labels = [
        "SI.NO", "Model No", "Item Description", "Unit",
        "Serial/Batch Number", "Warehouse Name", "Warehouse Location", "Remarks",
    ]
    for col, label in enumerate(col_labels, start=1):
        c = ws.cell(row=header_row, column=col, value=label)
        c.font = col_header_font
        c.fill = col_header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER_ALL
    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = f"A{header_row + 1}"

    # ── data rows (banded per line item, hand-fill columns tinted) ─────────
    r = header_row + 1
    for line_idx, (_, line) in enumerate(lines_df.iterrows()):
        article_number, description = item_lookup.get(line.get("product_id"), (None, None))
        quantity = int(line["quantity"])
        band = band_fill if line_idx % 2 == 0 else None
        for k in range(1, quantity + 1):
            values = [
                int(line["line_no"]), article_number or "", description or "",
                f"{k} of {quantity}", "", "", "", "",
            ]
            for col, val in enumerate(values, start=1):
                c = ws.cell(row=r, column=col, value=val)
                c.border = BORDER_ALL
                c.font = value_font
                if col in (1, 4):
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(col in (3, 8)))
                if col in (5, 6, 7, 8):
                    c.fill = handfill_fill
                elif band is not None:
                    c.fill = band
            r += 1

    # ── footer note ──────────────────────────────────────────────────────────
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=_PICK_SHEET_COLUMNS)
    note = ws.cell(
        row=r + 1, column=1,
        value="Fields shaded in yellow (Serial/Batch Number, Warehouse Name, Warehouse Location, Remarks) "
              "are completed by the Warehouse Team.",
    )
    note.font = footer_font

    ws.sheet_view.showGridLines = False

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Manager-insights aggregates (rep / customer sample stats)
# ─────────────────────────────────────────────────────────────────────────────

def _sample_request_item_summary(sample_request_id: int) -> tuple[int, str]:
    """
    Per-request total quantity and a comma-joined item summary (e.g.
    "Model A x2, Model B x1") for a single sample request's lines, built in
    Python from a per-request line query rather than SQL string
    aggregation, as specified in the Task 3 brief.
    """
    lines_df = query_df(
        """
        SELECT i.article_number, srl.quantity
        FROM sample_request_lines srl
        JOIN items i ON i.product_id = srl.product_id
        WHERE srl.sample_request_id = :sid
        ORDER BY srl.line_no
        """,
        {"sid": sample_request_id},
    )
    if lines_df.empty:
        return 0, ""
    total_qty = int(lines_df["quantity"].sum())
    item_summary = ", ".join(
        f"{_norm(row['article_number']) or '—'} x{int(row['quantity'])}"
        for _, row in lines_df.iterrows()
    )
    return total_qty, item_summary


def get_rep_sample_stats(rep_user_id: int) -> dict:
    """
    Manager-insights aggregate for a single rep: total sample quantity
    handed off (status DONE) all-time and so far this calendar month
    (bucketed by coordinator_done_at, i.e. when the coordinator actually
    marked the request Done, not when it was submitted/approved), plus
    the rep's 5 most-recent DONE requests. A rep with zero DONE history
    returns an all-zero/empty shape, not an error.
    """
    all_time_qty = query_scalar(
        """
        SELECT COALESCE(SUM(srl.quantity), 0)
        FROM sample_requests sr
        JOIN sample_request_lines srl ON srl.sample_request_id = sr.sample_request_id
        WHERE sr.rep_user_id = :rep_user_id AND sr.status = 'DONE'
        """,
        {"rep_user_id": rep_user_id},
    )
    this_month_qty = query_scalar(
        """
        SELECT COALESCE(SUM(srl.quantity), 0)
        FROM sample_requests sr
        JOIN sample_request_lines srl ON srl.sample_request_id = sr.sample_request_id
        WHERE sr.rep_user_id = :rep_user_id AND sr.status = 'DONE'
          AND sr.coordinator_done_at >= date_trunc('month', CURRENT_DATE)
        """,
        {"rep_user_id": rep_user_id},
    )

    recent_df = query_df(
        """
        SELECT sr.sample_request_id, sr.request_number, sr.coordinator_done_at,
               c.account_name AS customer_name
        FROM sample_requests sr
        JOIN customers c ON c.customer_id = sr.customer_id
        WHERE sr.rep_user_id = :rep_user_id AND sr.status = 'DONE'
        ORDER BY sr.coordinator_done_at DESC
        LIMIT 5
        """,
        {"rep_user_id": rep_user_id},
    )

    recent = []
    for _, row in recent_df.iterrows():
        total_qty, item_summary = _sample_request_item_summary(int(row["sample_request_id"]))
        recent.append({
            "request_number": row["request_number"],
            "coordinator_done_at": row["coordinator_done_at"],
            "customer_name": row["customer_name"],
            "total_qty": total_qty,
            "item_summary": item_summary,
        })

    return {
        "all_time_qty": int(all_time_qty or 0),
        "this_month_qty": int(this_month_qty or 0),
        "recent": recent,
    }


def get_customer_sample_stats(customer_id: int) -> dict:
    """
    Same shape as get_rep_sample_stats, scoped by customer instead of rep.
    Deliberately has no rep_user_id filter — a customer can receive DONE
    samples from multiple different reps over time, and all of them should
    count toward this customer's totals (cross-rep aggregation). Each
    `recent` entry additionally carries rep_name so a manager can see who
    handled each historical request.
    """
    all_time_qty = query_scalar(
        """
        SELECT COALESCE(SUM(srl.quantity), 0)
        FROM sample_requests sr
        JOIN sample_request_lines srl ON srl.sample_request_id = sr.sample_request_id
        WHERE sr.customer_id = :customer_id AND sr.status = 'DONE'
        """,
        {"customer_id": customer_id},
    )
    this_month_qty = query_scalar(
        """
        SELECT COALESCE(SUM(srl.quantity), 0)
        FROM sample_requests sr
        JOIN sample_request_lines srl ON srl.sample_request_id = sr.sample_request_id
        WHERE sr.customer_id = :customer_id AND sr.status = 'DONE'
          AND sr.coordinator_done_at >= date_trunc('month', CURRENT_DATE)
        """,
        {"customer_id": customer_id},
    )

    recent_df = query_df(
        """
        SELECT sr.sample_request_id, sr.request_number, sr.coordinator_done_at,
               c.account_name AS customer_name, u.name AS rep_name
        FROM sample_requests sr
        JOIN customers c ON c.customer_id = sr.customer_id
        JOIN users u ON u.user_id = sr.rep_user_id
        WHERE sr.customer_id = :customer_id AND sr.status = 'DONE'
        ORDER BY sr.coordinator_done_at DESC
        LIMIT 5
        """,
        {"customer_id": customer_id},
    )

    recent = []
    for _, row in recent_df.iterrows():
        total_qty, item_summary = _sample_request_item_summary(int(row["sample_request_id"]))
        recent.append({
            "request_number": row["request_number"],
            "coordinator_done_at": row["coordinator_done_at"],
            "customer_name": row["customer_name"],
            "rep_name": row["rep_name"],
            "total_qty": total_qty,
            "item_summary": item_summary,
        })

    return {
        "all_time_qty": int(all_time_qty or 0),
        "this_month_qty": int(this_month_qty or 0),
        "recent": recent,
    }
