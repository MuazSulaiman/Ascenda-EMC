# tests/test_sample_pick_sheet_export.py
"""
Coverage for app_pages.sample_request_helpers.generate_sample_pick_sheet
(Task 3). Requires DATABASE_URL env var and runs against the live dev DB,
no mocking — mirrors the fixture style of
tests/test_sample_request_helpers_smoke.py: a TEST_MARKER constant,
self-cleaning fixtures (yield + teardown), and direct DB-state / workbook
assertions.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import io
import uuid
from datetime import date

import pytest
from openpyxl import load_workbook
from sqlalchemy import text

from app_pages.sample_request_helpers import (
    submit_sample_request,
    manager_approve,
    generate_sample_pick_sheet,
)
from db import engine
from db_ops import query_scalar, exec_sql

TEST_MARKER = "PYTEST_SAMPLE_PICK_SHEET_TMP"


# ─────────────────────────────────────────────────────────────────────────────
# Fixtures (same pattern as test_sample_request_helpers_smoke.py)
# ─────────────────────────────────────────────────────────────────────────────

@pytest.fixture
def rep_user_id():
    uid = query_scalar("SELECT user_id FROM users WHERE role = 'rep' ORDER BY user_id LIMIT 1")
    assert uid is not None, "Need at least one rep user in DB"
    return int(uid)


@pytest.fixture
def any_customer_id():
    cid = query_scalar("SELECT customer_id FROM customers ORDER BY customer_id LIMIT 1")
    assert cid is not None, "Need at least one customer in DB"
    return int(cid)


@pytest.fixture
def product_ids():
    """Two distinct product_ids, needed for a multi-line pick-sheet request."""
    df_ids = query_scalar("SELECT COUNT(DISTINCT product_id) FROM items")
    assert df_ids is not None and df_ids >= 2, "Need at least two items in DB"
    from db_ops import query_df
    df = query_df("SELECT product_id FROM items ORDER BY product_id LIMIT 3")
    ids = df["product_id"].tolist()
    assert len(ids) >= 2, "Need at least two items in DB"
    return ids


def _delete_test_user(uid: int):
    exec_sql(
        """
        DELETE FROM sample_requests
        WHERE rep_user_id = :uid OR manager_user_id = :uid
           OR coordinator_user_id = :uid OR submitted_by = :uid OR withdrawn_by = :uid
        """,
        {"uid": uid},
    )
    exec_sql(
        "DELETE FROM notifications WHERE recipient_user_id = :uid OR actor_user_id = :uid",
        {"uid": uid},
    )
    exec_sql("DELETE FROM users WHERE user_id = :uid", {"uid": uid})


def _create_test_user(role: str, label: str) -> int:
    email = f"pytest.{label}.{uuid.uuid4().hex[:10]}@example.test"
    with engine.begin() as conn:
        uid = conn.execute(
            text("""
                INSERT INTO users (email, password_hash, name, role, is_active)
                VALUES (:email, 'pytest-not-a-real-hash', :name, :role, TRUE)
                RETURNING user_id
            """),
            {"email": email, "name": f"{TEST_MARKER} {label}", "role": role},
        ).scalar_one()
    return int(uid)


@pytest.fixture
def sales_manager_id():
    uid = _create_test_user("sales manager", "mgr")
    yield uid
    _delete_test_user(uid)


@pytest.fixture
def sids():
    ids: list[int] = []
    yield ids
    for sid in ids:
        exec_sql(
            "DELETE FROM notifications WHERE (link_params->>'sample_request_id')::int = :sid",
            {"sid": sid},
        )
        exec_sql("DELETE FROM sample_requests WHERE sample_request_id = :sid", {"sid": sid})


def _base_header(customer_id: int, **overrides) -> dict:
    header = {"customer_id": customer_id, "request_date": date.today(), "remarks": TEST_MARKER}
    header.update(overrides)
    return header


# ─────────────────────────────────────────────────────────────────────────────
# Tests
# ─────────────────────────────────────────────────────────────────────────────

def test_pick_sheet_multiline_quantities_produce_exploded_rows(
    sids, rep_user_id, sales_manager_id, any_customer_id, product_ids
):
    """
    A 3-line request with quantities [2, 1, 5] must explode into exactly
    2 + 1 + 5 = 8 data rows, with the Unit column reading "1 of 2".."2 of 2"
    for the first line (qty 2), "1 of 1" for the second (qty 1), and
    "1 of 5".."5 of 5" for the third (qty 5).
    """
    p1, p2, p3 = (product_ids * 2)[:3]  # tolerate fewer than 3 distinct ids by cycling
    lines = [
        {"product_id": p1, "quantity": 2},
        {"product_id": p2, "quantity": 1},
        {"product_id": p3, "quantity": 5},
    ]
    header = _base_header(any_customer_id, delivery_date=date.today())
    sid, snum = submit_sample_request(header, lines, rep_user_id)
    sids.append(sid)

    ok, err = manager_approve(sid, sales_manager_id)
    assert ok, err

    xlsx_bytes = generate_sample_pick_sheet(sid)
    assert isinstance(xlsx_bytes, bytes)
    assert len(xlsx_bytes) > 0

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.sheetnames == ["Warehouse Pick Sheet"]
    ws = wb.active

    all_rows = list(ws.iter_rows(values_only=True))
    # Row 1: Request No, Row 2: Date, Row 3: Delivery Date, Row 4: Customer,
    # Row 5: blank, Row 6: column headers.
    assert all_rows[0][0] == f"Request No: {snum}"
    assert str(all_rows[1][0]).startswith("Date:")
    assert str(all_rows[2][0]).startswith("Delivery Date:")
    assert str(all_rows[3][0]).startswith("Customer:")
    assert all_rows[4] == (None,) or all(v is None for v in all_rows[4])
    assert list(all_rows[5][:7]) == [
        "SI.NO", "Model No", "Item Description", "Unit",
        "Serial/Batch Number", "Warehouse Name", "Warehouse Location",
    ]

    data_rows = all_rows[6:]
    assert len(data_rows) == 8, f"expected 8 exploded rows (2+1+5), got {len(data_rows)}"

    line1_rows = data_rows[0:2]
    line2_rows = data_rows[2:3]
    line3_rows = data_rows[3:8]

    assert [r[3] for r in line1_rows] == ["1 of 2", "2 of 2"]
    assert all(r[0] == 1 for r in line1_rows)

    assert [r[3] for r in line2_rows] == ["1 of 1"]
    assert all(r[0] == 2 for r in line2_rows)

    assert [r[3] for r in line3_rows] == ["1 of 5", "2 of 5", "3 of 5", "4 of 5", "5 of 5"]
    assert all(r[0] == 3 for r in line3_rows)

    # Last three columns (serial/batch, warehouse name, warehouse location) blank for hand-filling.
    for r in data_rows:
        assert r[4] in (None, ""), r
        assert r[5] in (None, ""), r
        assert r[6] in (None, ""), r


def test_pick_sheet_rejects_non_approved_status(sids, rep_user_id, any_customer_id, product_ids):
    """Calling generate_sample_pick_sheet on an IN_REVIEW request must raise ValueError."""
    lines = [{"product_id": product_ids[0], "quantity": 1}]
    sid, _ = submit_sample_request(_base_header(any_customer_id), lines, rep_user_id)
    sids.append(sid)

    header_status = query_scalar("SELECT status FROM sample_requests WHERE sample_request_id = :sid", {"sid": sid})
    assert header_status == "IN_REVIEW"

    with pytest.raises(ValueError):
        generate_sample_pick_sheet(sid)


def test_pick_sheet_missing_request_raises_value_error():
    with pytest.raises(ValueError):
        generate_sample_pick_sheet(-999999)
